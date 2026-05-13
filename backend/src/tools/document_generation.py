# =============================================================================
# PH Agent Hub — Document Generation Tool Factory
# =============================================================================
# Markdown → PDF (weasyprint), list-of-dicts → Excel (openpyxl),
# CSV export. Output artifacts to MinIO/S3.
#
# Dependencies: weasyprint, markdown, openpyxl (already installed)
# =============================================================================

import csv
import io
import logging
import uuid
from typing import Any

from agent_framework import tool

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
DEFAULT_PRESIGNED_TTL: int = 3600  # 1 hour


def _get_bucket_prefix() -> str:
    """Return the MinIO bucket prefix from settings."""
    from ..core.config import settings
    return settings.MINIO_BUCKET_PREFIX


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _get_bucket(tenant_id: str) -> str:
    """Return the MinIO bucket name for a tenant."""
    return f"{_get_bucket_prefix()}-{tenant_id}"


async def _upload_and_get_url(
    bucket: str,
    key: str,
    data: bytes,
    content_type: str,
    expires_in: int = DEFAULT_PRESIGNED_TTL,
) -> str:
    """Upload data to MinIO/S3 and return a presigned download URL."""
    from ..storage.s3 import generate_presigned_url, upload_object
    await upload_object(bucket, key, data, content_type)
    return await generate_presigned_url(bucket, key, expires_in=expires_in)


# ---------------------------------------------------------------------------
# Tool factory
# ---------------------------------------------------------------------------


def build_document_generation_tools(
    tool_config: dict | None = None,
    tenant_id: str = "",
) -> list:
    """Return a list of MAF @tool-decorated async functions for document
    generation.

    Args:
        tool_config: Optional ``Tool.config`` JSON dict.  May include:
            - ``company_logo_url`` (str): URL of company logo for PDF header
            - ``default_format`` (str): default output format ("pdf", "excel", "csv")
        tenant_id: The tenant ID for MinIO bucket resolution.

    Returns:
        A list of callables ready to pass to ``Agent(tools=...)``.
    """
    config = tool_config or {}
    company_logo_url: str = config.get("company_logo_url", "")

    @tool
    async def generate_pdf(markdown: str, title: str = "Report") -> dict:
        """Generate a PDF document from Markdown content.

        The Markdown is converted to HTML and then rendered as a PDF
        using WeasyPrint. The resulting PDF is stored and a download
        URL is returned.

        Args:
            markdown: The Markdown content to convert to PDF.
            title: The document title (used in PDF metadata).

        Returns:
            A dict with:
            - ``url``: presigned download URL for the generated PDF
            - ``filename``: suggested filename for the PDF
            - ``size_bytes``: size of the generated PDF
            - ``error``: error message if generation failed
        """
        if not markdown or not markdown.strip():
            return {"error": "No markdown content provided to generate PDF"}

        try:
            import markdown as md_lib
            html_content = md_lib.markdown(
                markdown,
                extensions=["tables", "fenced_code", "codehilite", "toc", "nl2br"],
            )
        except ImportError:
            # Fallback if markdown lib is not available
            html_content = markdown.replace("\n", "<br>\n")

        # Build a complete HTML document with basic styling
        logo_html = ""
        if company_logo_url:
            logo_html = f'<img src="{company_logo_url}" alt="Logo" style="max-height:60px; margin-bottom:20px;">'

        html_doc = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{title}</title>
<style>
  @page {{ margin: 2cm; size: A4; }}
  body {{ font-family: 'DejaVu Sans', Arial, sans-serif; font-size: 11pt; line-height: 1.6; color: #333; }}
  h1 {{ font-size: 20pt; margin-top: 0; color: #1a1a1a; }}
  h2 {{ font-size: 16pt; color: #2a2a2a; }}
  h3 {{ font-size: 13pt; color: #3a3a3a; }}
  table {{ border-collapse: collapse; width: 100%; margin: 1em 0; }}
  th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
  th {{ background-color: #f5f5f5; font-weight: bold; }}
  code {{ background-color: #f0f0f0; padding: 2px 4px; border-radius: 3px; font-size: 9pt; }}
  pre {{ background-color: #f5f5f5; padding: 12px; border-radius: 4px; overflow-x: auto; font-size: 9pt; }}
  pre code {{ background: none; padding: 0; }}
  .header {{ border-bottom: 2px solid #1a1a1a; padding-bottom: 10px; margin-bottom: 20px; }}
  .footer {{ border-top: 1px solid #ccc; padding-top: 10px; margin-top: 30px; font-size: 9pt; color: #999; }}
</style>
</head>
<body>
{logo_html}
{html_content}
<div class="footer">Generated by PH Agent Hub</div>
</body>
</html>"""

        try:
            from weasyprint import HTML
            pdf_bytes = HTML(string=html_doc).write_pdf()
        except ImportError:
            return {
                "error": (
                    "PDF generation is not available. The weasyprint library "
                    "is not installed. Please contact your administrator."
                )
            }
        except Exception as exc:
            logger.error("WeasyPrint PDF generation failed: %s", exc)
            return {"error": f"PDF generation failed: {str(exc)}"}

        # Upload to MinIO/S3
        if not tenant_id:
            return {
                "error": "Tenant ID not available for file storage",
                "pdf_base64": None,
            }

        try:
            bucket = _get_bucket(tenant_id)
            file_id = str(uuid.uuid4())
            key = f"generated/{file_id}.pdf"
            url = await _upload_and_get_url(
                bucket, key, pdf_bytes, "application/pdf"
            )
            safe_title = "".join(c for c in title if c.isalnum() or c in " _-")[:50]
            filename = f"{safe_title}.pdf" if safe_title else "report.pdf"

            logger.info("Generated PDF: %s (%d bytes)", key, len(pdf_bytes))
            return {
                "url": url,
                "filename": filename,
                "size_bytes": len(pdf_bytes),
            }
        except Exception as exc:
            logger.error("Failed to upload generated PDF: %s", exc)
            return {"error": f"Failed to store generated PDF: {str(exc)}"}

    # ------------------------------------------------------------------
    @tool
    async def generate_excel(data: list[dict], sheet_name: str = "Sheet1") -> dict:
        """Generate an Excel (.xlsx) spreadsheet from a list of dictionaries.

        Each dictionary in the list becomes a row, keys become column headers.

        Args:
            data: A list of dicts where each dict represents a row.
                  Example: [{"Name": "Alice", "Age": 30}, {"Name": "Bob", "Age": 25}]
            sheet_name: Name of the worksheet (default "Sheet1").

        Returns:
            A dict with:
            - ``url``: presigned download URL for the generated Excel file
            - ``filename``: suggested filename
            - ``size_bytes``: size of the generated file
            - ``row_count``: number of data rows written
            - ``error``: error message if generation failed
        """
        if not data:
            return {"error": "No data provided to generate Excel file"}

        if not isinstance(data, list):
            return {"error": "Data must be a list of dictionaries"}

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name[:31]  # Excel sheet name limit

            # Extract headers from the first dict
            if data and isinstance(data[0], dict):
                headers = list(data[0].keys())
            else:
                # Try to infer headers
                all_keys = set()
                for item in data:
                    if isinstance(item, dict):
                        all_keys.update(item.keys())
                headers = sorted(all_keys)

            # Style for headers
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=str(header))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Write data rows
            row_count = 0
            for row_idx, item in enumerate(data, 2):
                if not isinstance(item, dict):
                    continue
                for col_idx, header in enumerate(headers, 1):
                    val = item.get(header, "")
                    # Convert non-serializable types
                    if isinstance(val, (list, dict)):
                        val = str(val)
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.border = thin_border
                row_count += 1

            # Auto-fit column widths (approximate)
            for col_idx, header in enumerate(headers, 1):
                max_width = len(str(header)) + 2
                for row_idx in range(2, row_count + 2):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val:
                        max_width = max(max_width, min(len(str(cell_val)) + 2, 50))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_width

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            excel_bytes = output.getvalue()
            output.close()

        except ImportError:
            return {
                "error": (
                    "Excel generation is not available. The openpyxl library "
                    "is not installed. Please contact your administrator."
                )
            }
        except Exception as exc:
            logger.error("Excel generation failed: %s", exc)
            return {"error": f"Excel generation failed: {str(exc)}"}

        # Upload to MinIO/S3
        if not tenant_id:
            return {"error": "Tenant ID not available for file storage"}

        try:
            bucket = _get_bucket(tenant_id)
            file_id = str(uuid.uuid4())
            safe_name = "".join(c for c in sheet_name if c.isalnum() or c in " _-")[:31]
            filename = f"{safe_name}.xlsx" if safe_name else "data.xlsx"
            key = f"generated/{file_id}.xlsx"
            url = await _upload_and_get_url(
                bucket, key, excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            logger.info("Generated Excel: %s (%d bytes, %d rows)", key, len(excel_bytes), row_count)
            return {
                "url": url,
                "filename": filename,
                "size_bytes": len(excel_bytes),
                "row_count": row_count,
            }
        except Exception as exc:
            logger.error("Failed to upload generated Excel: %s", exc)
            return {"error": f"Failed to store generated Excel: {str(exc)}"}

    # ------------------------------------------------------------------
    @tool
    async def generate_csv(data: list[dict]) -> dict:
        """Generate a CSV file from a list of dictionaries.

        Each dictionary in the list becomes a row, keys become column headers.

        Args:
            data: A list of dicts where each dict represents a row.
                  Example: [{"Name": "Alice", "Age": 30}, {"Name": "Bob", "Age": 25}]

        Returns:
            A dict with:
            - ``url``: presigned download URL for the generated CSV file
            - ``filename``: suggested filename
            - ``size_bytes``: size of the generated file
            - ``row_count``: number of data rows written
            - ``error``: error message if generation failed
        """
        if not data:
            return {"error": "No data provided to generate CSV file"}

        if not isinstance(data, list):
            return {"error": "Data must be a list of dictionaries"}

        try:
            # Extract headers
            if data and isinstance(data[0], dict):
                headers = list(data[0].keys())
            else:
                all_keys = set()
                for item in data:
                    if isinstance(item, dict):
                        all_keys.update(item.keys())
                headers = sorted(all_keys)

            output = io.StringIO(newline="")
            writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
            writer.writeheader()

            row_count = 0
            for item in data:
                if isinstance(item, dict):
                    # Convert non-serializable values to strings
                    clean_item = {}
                    for k, v in item.items():
                        if isinstance(v, (list, dict)):
                            clean_item[k] = str(v)
                        elif v is None:
                            clean_item[k] = ""
                        else:
                            clean_item[k] = v
                    writer.writerow(clean_item)
                    row_count += 1

            csv_bytes = output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
            output.close()

        except Exception as exc:
            logger.error("CSV generation failed: %s", exc)
            return {"error": f"CSV generation failed: {str(exc)}"}

        # Upload to MinIO/S3
        if not tenant_id:
            return {"error": "Tenant ID not available for file storage"}

        try:
            bucket = _get_bucket(tenant_id)
            file_id = str(uuid.uuid4())
            key = f"generated/{file_id}.csv"
            url = await _upload_and_get_url(
                bucket, key, csv_bytes, "text/csv"
            )

            logger.info("Generated CSV: %s (%d bytes, %d rows)", key, len(csv_bytes), row_count)
            return {
                "url": url,
                "filename": "export.csv",
                "size_bytes": len(csv_bytes),
                "row_count": row_count,
            }
        except Exception as exc:
            logger.error("Failed to upload generated CSV: %s", exc)
            return {"error": f"Failed to store generated CSV: {str(exc)}"}

    return [generate_pdf, generate_excel, generate_csv]
